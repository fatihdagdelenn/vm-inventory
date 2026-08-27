"""
Reporting service: Excel (.xlsx), CSV and PDF export.
Filtered search results export as-is (the search parameter is reapplied).
"""
import csv
import io
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
import logging

from ..core.timezone import now_local

logger = logging.getLogger("report")

# --- PDF Turkish font registration ---------------------------------------
# Default Helvetica cannot render Turkish letters (Latin-5) like ş/ğ/İ/ı;
# they come out broken/boxed in PDFs. A Unicode DejaVuSans ships in the repo
# and is registered here. If missing, we fall back to Helvetica (the PDF is
# still produced, only Turkish letters may be off).
_FONT_DIR = os.path.join(os.path.dirname(__file__), "..", "static", "fonts")
PDF_FONT = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"
try:
    _regular = os.path.join(_FONT_DIR, "DejaVuSans.ttf")
    _bold = os.path.join(_FONT_DIR, "DejaVuSans-Bold.ttf")
    if os.path.exists(_regular) and os.path.exists(_bold):
        pdfmetrics.registerFont(TTFont("DejaVuSans", _regular))
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", _bold))
        pdfmetrics.registerFontFamily(
            "DejaVuSans", normal="DejaVuSans", bold="DejaVuSans-Bold")
        PDF_FONT, PDF_FONT_BOLD = "DejaVuSans", "DejaVuSans-Bold"
    else:
        logger.warning("DejaVuSans font not found (%s); Turkish characters "
                       "may be missing in PDFs.", _FONT_DIR)
except Exception as exc:  # pragma: no cover - font registration is non-critical
    logger.warning("PDF fontu kaydedilemedi: %s", exc)

# Exported VM columns: (header, model field)
VM_COLUMNS = [
    ("VM Adı", "name"), ("VM ID", "vmid"), ("IP Adresleri", "ip_addresses"),
    ("MAC Adresleri", "mac_addresses"), ("İşletim Sistemi", "guest_os"),
    ("CPU (adet)", "cpu_count"), ("CPU Kullanım (%)", "cpu_usage_pct"),
    ("RAM Tahsis (MB)", "ram_mb"), ("RAM Kullanım (MB)", "ram_usage_mb"),
    ("Disk Tahsis (GB)", "disk_total_gb"), ("Disk Kullanım (GB)", "disk_used_gb"),
    ("Güç Durumu", "power_state"), ("Host", "host_name"), ("Cluster", "cluster"),
    ("Datastore", "datastore"), ("VLAN", "vlans"), ("Ortam", "environment"),
    ("Sahip", "owner"), ("Tools/Agent", "tools_status"),
    ("Platform Notu", "guest_notes"),
]

# Physical inventory columns. device_type/status/role resolve to Turkish
# labels via _row_values (the values stored in DB are codes).
PHYSICAL_DTYPE_LABELS = {"server": "Fiziksel Sunucu", "storage": "Storage",
                         "san_switch": "SAN Switch", "backup": "Yedekleme Ünitesi"}
PHYSICAL_STATUS_LABELS = {"active": "Aktif", "passive": "Pasif", "faulty": "Arızalı",
                          "retired": "Emekli", "spare": "Yedek"}
PHYSICAL_ROLE_LABELS = {"hypervisor": "Hypervisor", "windows": "Windows",
                        "linux": "Linux", "other": "Diğer"}
PHYSICAL_COLUMNS = [
    ("Tip", "device_type"), ("Ad", "name"), ("Rol", "role"),
    ("Lokasyon", "location"), ("Durum", "status"), ("Yönetim IP", "mgmt_ip"),
    ("iLO/BMC IP", "ilo_ip"), ("Marka", "brand"), ("Model", "model"),
    ("Seri No", "serial_no"), ("CPU", "cpu"), ("RAM (GB)", "ram_gb"),
    ("İşletim Sistemi", "os"), ("Not", "notes"),
]

HOST_COLUMNS = [
    ("Host Adı", "name"), ("Yönetim IP", "mgmt_ip"), ("İşletim Sistemi", "os_version"),
    ("CPU Modeli", "cpu_model"), ("Çekirdek", "cpu_cores"),
    ("Toplam RAM (MB)", "ram_total_mb"), ("Kullanılan RAM (MB)", "ram_used_mb"),
    ("Disk (GB)", "disk_total_gb"), ("Cluster", "cluster"), ("Durum", "status"),
]

DATASTORE_COLUMNS = [
    ("Datastore", "name"), ("Node", "node"), ("Platform", "platform_name"),
    ("Tip", "type"), ("Kapasite (GB)", "capacity_gb"), ("Kullanılan (GB)", "used_gb"),
    ("Boş (GB)", "free_gb"), ("Kullanım %", "usage_pct"),
    ("Host", "host_count"), ("VM", "vm_count"), ("Durum", "status"),
]

SNAPSHOT_COLUMNS = [
    ("VM", "vm_name"), ("Snapshot", "name"), ("Platform", "platform_name"),
    ("Oluşturma", "created_at"), ("Yaş (gün)", "age_days"),
    ("Üst Snapshot", "parent"), ("Açıklama", "description"),
]

BACKUP_COLUMNS = [
    ("VM", "vm_name"), ("VMID", "vmid"), ("Depo", "storage"), ("Kaynak", "source"),
    ("Format", "fmt"), ("Oluşturma", "created_at"), ("Yaş (gün)", "age_days"),
    ("Boyut (GB)", "size_gb"), ("Korumalı", "protected"), ("Not", "notes"),
]


def _row_values(obj, columns):
    """Extract values from a model object OR a dict, in column order.

    Physical host projections are passed as dicts; manual devices and other
    inventory are ORM objects. device_type/status/role codes -> TR labels.
    """
    is_dict = isinstance(obj, dict)
    get = (lambda f: obj.get(f)) if is_dict else (lambda f: getattr(obj, f, ""))
    is_physical = is_dict or obj.__class__.__name__ == "PhysicalDevice"
    values = []
    for _, field in columns:
        if field == "host_name":  # VM -> related host name
            values.append(obj.host_ref.name if getattr(obj, "host_ref", None) else "")
        elif field == "device_type":
            values.append(PHYSICAL_DTYPE_LABELS.get(get(field), get(field) or ""))
        elif field == "status" and is_physical:
            values.append(PHYSICAL_STATUS_LABELS.get(get(field), get(field) or ""))
        elif field == "role" and is_physical:
            values.append(PHYSICAL_ROLE_LABELS.get(get(field), get(field) or ""))
        elif field in ("cpu_usage_pct", "disk_used_gb", "disk_total_gb"):
            v = get(field)
            values.append("" if v is None else round(float(v), 1))
        elif field.startswith("disk_size_"):
            # Per-disk allocated size: disk_size_1, disk_size_2, ...
            idx = int(field.rsplit("_", 1)[1]) - 1
            raw = getattr(obj, "disks_json", None) if not is_dict else obj.get("disks_json")
            sz = ""
            if raw:
                try:
                    dl = json.loads(raw)
                    if 0 <= idx < len(dl):
                        s = dl[idx].get("size_gb")
                        sz = "" if s is None else round(float(s), 1)
                except (ValueError, TypeError, AttributeError, IndexError):
                    pass
            values.append(sz)
        else:
            v = get(field)
            values.append("" if v is None else v)
    return values


def _max_disk_count(items):
    """Largest disk count across VM items (for dynamic per-disk columns)."""
    n = 0
    for obj in items:
        raw = obj.get("disks_json") if isinstance(obj, dict) else getattr(obj, "disks_json", None)
        if raw:
            try:
                n = max(n, len(json.loads(raw)))
            except (ValueError, TypeError):
                pass
    return n


def vm_columns_with_disks(items):
    """VM_COLUMNS plus one 'Disk N (GB)' column per disk, up to the max disk
    count in the data. Single-disk fleets get no extra columns (max=1)."""
    n = _max_disk_count(items)
    if n <= 1:
        return VM_COLUMNS
    extra = [(f"Disk {i} (GB)", f"disk_size_{i}") for i in range(1, n + 1)]
    # Insert the per-disk columns right after "Disk Kullanım (GB)"
    out, injected = [], False
    for col in VM_COLUMNS:
        out.append(col)
        if col[1] == "disk_used_gb":
            out.extend(extra); injected = True
    if not injected:
        out.extend(extra)
    return out


def export_excel(items, columns, title="Envanter Raporu") -> bytes:
    """Produce a formatted Excel report."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_fill = PatternFill("solid", fgColor="1B3A57")
    header_font = Font(color="FFFFFF", bold=True)
    for col, (header, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal="center")

    for row, item in enumerate(items, 2):
        for col, value in enumerate(_row_values(item, columns), 1):
            ws.cell(row=row, column=col, value=value)

    # Fit column widths to the content
    for col in range(1, len(columns) + 1):
        max_len = max((len(str(ws.cell(row=r, column=col).value or ""))
                       for r in range(1, min(ws.max_row, 200) + 1)), default=10)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 45)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_excel_multi(sections, title="Envanter Raporu") -> bytes:
    """Multi-sheet Excel: one worksheet per (sheet_name, items, columns) section.

    Used for the combined "all inventory" report so VMs, hosts, datastores and
    physical devices land in a single workbook on separate tabs.
    """
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1B3A57")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet_name, items, columns in sections:
        ws = wb.create_sheet(title=sheet_name[:31])
        for col, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill, cell.font = header_fill, header_font
            cell.alignment = Alignment(horizontal="center")
        for row, item in enumerate(items, 2):
            for col, value in enumerate(_row_values(item, columns), 1):
                ws.cell(row=row, column=col, value=value)
        for col in range(1, len(columns) + 1):
            max_len = max((len(str(ws.cell(row=r, column=col).value or ""))
                           for r in range(1, min(ws.max_row, 200) + 1)), default=10)
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 45)
        ws.freeze_panes = "A2"
        if ws.max_row >= 1:
            ws.auto_filter.ref = ws.dimensions
    if not wb.sheetnames:            # no sections -> keep a valid empty book
        wb.create_sheet(title="Bos")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_csv(items, columns) -> bytes:
    """CSV with a UTF-8 BOM (Turkish character compatibility in Excel)."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow([h for h, _ in columns])
    for item in items:
        writer.writerow(_row_values(item, columns))
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def export_pdf(items, columns, title="Envanter Raporu") -> bytes:
    """
        Landscape A4 PDF table report. Column parity with Excel/CSV: ALL columns
        are included (only the first 10 used to be).
        """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=8*mm, rightMargin=8*mm,
                            topMargin=12*mm, bottomMargin=12*mm)

    base = getSampleStyleSheet()
    title_style = ParagraphStyle("trTitle", parent=base["Title"],
                                 fontName=PDF_FONT_BOLD)
    meta_style = ParagraphStyle("trMeta", parent=base["Normal"],
                                fontName=PDF_FONT, fontSize=9)
    cell_style = ParagraphStyle("trCell", parent=base["Normal"],
                                fontName=PDF_FONT, fontSize=6, leading=7)
    head_style = ParagraphStyle("trHead", parent=base["Normal"],
                                fontName=PDF_FONT_BOLD, fontSize=6, leading=7,
                                textColor=colors.white)

    elements = [
        Paragraph(title, title_style),
        Paragraph(f"Oluşturulma: {now_local():%d.%m.%Y %H:%M} — Kayıt: {len(items)}",
                  meta_style),
        Spacer(1, 5*mm),
    ]

    # All columns - wrap cells in Paragraph so long text wraps to the next line
    data = [[Paragraph(str(h), head_style) for h, _ in columns]]
    for item in items:
        data.append([Paragraph(_pdf_escape(v), cell_style)
                     for v in _row_values(item, columns)])

    # Split column widths evenly across the page (total = usable width)
    n = len(columns)
    col_w = (doc.width / n) if n else doc.width
    table = Table(data, colWidths=[col_w] * n, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B3A57")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F5F8")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def _pdf_escape(value) -> str:
    """Safe text for Paragraph: None->'', escape XML special characters."""
    s = "" if value is None else str(value)
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def export_csv_multi(sections) -> bytes:
    """Combined CSV: each section prefixed with a '# <name>' banner row."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    for i, (name, items, columns) in enumerate(sections):
        if i:
            writer.writerow([])
        writer.writerow([f"# {name}"])
        writer.writerow([h for h, _ in columns])
        for item in items:
            writer.writerow(_row_values(item, columns))
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def export_pdf_multi(sections, title="Tüm Envanter") -> bytes:
    """Combined PDF: one titled table per section, stacked on landscape A4."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=8*mm, rightMargin=8*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    base = getSampleStyleSheet()
    title_style = ParagraphStyle("mTitle", parent=base["Title"],
                                 fontName=PDF_FONT_BOLD)
    sect_style = ParagraphStyle("mSect", parent=base["Heading2"],
                                fontName=PDF_FONT_BOLD, fontSize=12, spaceBefore=10)
    cell_style = ParagraphStyle("mCell", parent=base["Normal"],
                                fontName=PDF_FONT, fontSize=6, leading=7)
    head_style = ParagraphStyle("mHead", parent=base["Normal"],
                                fontName=PDF_FONT_BOLD, fontSize=6, leading=7,
                                textColor=colors.white)
    elements = [Paragraph(_pdf_escape(title), title_style), Spacer(1, 6)]
    for name, items, columns in sections:
        elements.append(Paragraph(f"{_pdf_escape(name)} ({len(items)})", sect_style))
        data = [[Paragraph(str(h), head_style) for h, _ in columns]]
        for item in items:
            data.append([Paragraph(_pdf_escape(v), cell_style)
                         for v in _row_values(item, columns)])
        n = len(columns)
        col_w = (doc.width / n) if n else doc.width
        table = Table(data, colWidths=[col_w] * n, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B3A57")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F5F8")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 1.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 8))
    doc.build(elements)
    return buf.getvalue()
